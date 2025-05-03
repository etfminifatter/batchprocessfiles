import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from utils.file_utils import create_dirs
import os
from tkinter import scrolledtext
import logging
from datetime import datetime
import re

class CreateDirsTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        # 初始化日志
        self.logger = logging.getLogger("create_dirs_tab")
        self.logger.info("初始化创建目录标签页")
        
        # 保存对root窗口的引用，用于访问剪贴板
        self.root = self.winfo_toplevel()
        self.setup_ui()
        
        # 设置默认目标路径
        default_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.target_path.set(default_path)
        
        # 添加默认内容示例
        self.text_input.insert("1.0", "# 请在此输入目录名称(每行一个)\ndir1\ndir2\ndir3")
        
        self.logger.info("创建目录标签页初始化完成")
        
    def setup_ui(self):
        """设置用户界面"""
        # 主框架
        main_frame = ttk.Frame(self, padding=(10, 10))
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 上部分：目录输入和设置
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 设置左右布局
        left_pane = ttk.Frame(top_frame, width=400)  # 设置固定宽度
        left_pane.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        left_pane.pack_propagate(False)  # 防止子组件改变frame大小
        
        # 右侧区域 - 不使用Canvas，回归简单Frame
        right_pane = ttk.Frame(top_frame, width=400)  # 设置固定宽度
        right_pane.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        right_pane.pack_propagate(False)  # 防止子组件改变frame大小
        
        # 设置各部分
        self.setup_input_area(left_pane)         # 左侧：输入区域
        self.setup_naming_rules(right_pane)      # 右上：命名规则
        self.setup_output_settings(right_pane)   # 右下：输出设置
        # 临时注释掉层级结构设置
        # self.setup_hierarchy_settings(right_pane) # 右中：层级结构设置
        
        # 预览区域 - 限制高度
        preview_frame = ttk.LabelFrame(main_frame, text="预览", height=150)
        preview_frame.pack(fill=tk.X, expand=False)
        preview_frame.pack_propagate(False)  # 防止子组件改变frame高度
        
        self.setup_preview_area(preview_frame)
        
        # 初始化界面状态
        self.toggle_input_method()
        self.toggle_naming_rule()
        # 临时注释掉层级结构切换
        # self.toggle_hierarchy()
        
        # 禁用层级结构功能
        self.enable_hierarchy = tk.BooleanVar(value=False)
    
    def setup_input_area(self, parent):
        """设置输入区域"""
        input_frame = ttk.LabelFrame(parent, text="输入")
        input_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 输入方式选择
        method_frame = ttk.Frame(input_frame)
        method_frame.pack(fill=tk.X, padx=10, pady=(8, 5))
        
        ttk.Label(method_frame, text="输入方式:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.input_method = tk.StringVar(value="direct")
        ttk.Radiobutton(method_frame, text="直接输入", variable=self.input_method, value="direct", 
                       command=self.toggle_input_method).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(method_frame, text="从文件导入", variable=self.input_method, value="file", 
                       command=self.toggle_input_method).pack(side=tk.LEFT, padx=0)
        
        # 直接输入框架
        self.direct_input_frame = ttk.Frame(input_frame)
        self.direct_input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 添加多行文本输入框，带滚动条
        self.text_input = scrolledtext.ScrolledText(self.direct_input_frame, height=10, width=40, wrap=tk.WORD)
        self.text_input.pack(fill=tk.BOTH, expand=True)
        
        # 按钮区域
        button_frame = ttk.Frame(self.direct_input_frame)
        button_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Button(button_frame, text="从剪贴板粘贴", style="Auxiliary.TButton", command=self.paste_from_clipboard).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="清空", style="Auxiliary.TButton", command=self.clear_input).pack(side=tk.LEFT, padx=0)
        
        # 从文件导入框架
        self.file_input_frame = ttk.Frame(input_frame)
        
        file_select_frame = ttk.Frame(self.file_input_frame)
        file_select_frame.pack(fill=tk.X, pady=5)
        
        self.file_path = tk.StringVar()
        ttk.Entry(file_select_frame, textvariable=self.file_path, width=30).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(file_select_frame, text="浏览", command=self.browse_file, style="Auxiliary.TButton").pack(side=tk.LEFT, padx=(5, 0))
        
        # 文件类型选择
        file_type_frame = ttk.Frame(self.file_input_frame)
        file_type_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(file_type_frame, text="文件类型:").pack(side=tk.LEFT)
        self.file_import_type = tk.StringVar(value=".txt")
        file_type_combo = ttk.Combobox(file_type_frame, textvariable=self.file_import_type, width=15)
        file_type_combo['values'] = ('.txt', '.csv', '.xlsx', '.xls')
        file_type_combo.pack(side=tk.LEFT, padx=(5, 0))
        # 监听文件类型变化
        file_type_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_file_preview())
        
        # 表头和列选择选项
        options_frame = ttk.Frame(self.file_input_frame)
        options_frame.pack(fill=tk.X, pady=5)
        
        self.file_has_header = tk.BooleanVar(value=False)
        header_checkbox = ttk.Checkbutton(options_frame, text="跳过表头行", variable=self.file_has_header, 
                                         command=self.refresh_file_preview)
        header_checkbox.pack(side=tk.LEFT)
        
        column_frame = ttk.Frame(options_frame)
        column_frame.pack(side=tk.LEFT, padx=(20, 0))
        
        ttk.Label(column_frame, text="列索引:").pack(side=tk.LEFT)
        self.column_index = tk.StringVar(value="1")
        column_entry = ttk.Entry(column_frame, textvariable=self.column_index, width=5)
        column_entry.pack(side=tk.LEFT, padx=(5, 0))
        # 绑定列索引变更事件
        column_entry.bind("<FocusOut>", lambda e: self.refresh_file_preview())
        self.column_index.trace_add("write", lambda *args: self.refresh_file_preview())
        
        # 文件预览区域
        ttk.Label(self.file_input_frame, text="文件内容预览:").pack(anchor=tk.W, pady=(5, 0))
        
        self.file_preview = scrolledtext.ScrolledText(self.file_input_frame, height=8, width=40, wrap=tk.WORD, state="disabled")
        self.file_preview.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 默认显示直接输入
        self.toggle_input_method()
    
    def setup_naming_rules(self, parent):
        """设置命名规则区域"""
        naming_frame = ttk.LabelFrame(parent, text="命名规则")
        naming_frame.pack(fill=tk.X, padx=5, pady=(0, 10))
        
        # 命名方式选择
        method_frame = ttk.Frame(naming_frame)
        method_frame.pack(fill=tk.X, padx=10, pady=(8, 5))
        
        ttk.Label(method_frame, text="命名方式:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.naming_rule = tk.StringVar(value="direct")
        ttk.Radiobutton(method_frame, text="直接命名", variable=self.naming_rule, value="direct", 
                       command=self.toggle_naming_rule).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(method_frame, text="自定义规则", variable=self.naming_rule, value="custom", 
                       command=self.toggle_naming_rule).pack(side=tk.LEFT, padx=0)
        
        # 直接命名框架
        self.direct_naming_frame = ttk.Frame(naming_frame)
        ttk.Label(self.direct_naming_frame, text="将直接使用输入的目录名").pack(pady=20)
        
        # 自定义命名规则框架 - 使用更紧凑的布局
        self.custom_rule_frame = ttk.Frame(naming_frame)
        
        # 使用更紧凑的布局，把规则说明和输入放在一起
        rule_frame = ttk.Frame(self.custom_rule_frame)
        rule_frame.pack(fill=tk.X, pady=5, padx=5)
        rule_frame.columnconfigure(1, weight=1)
        
        ttk.Label(rule_frame, text="命名规则:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.rule_pattern = ttk.Entry(rule_frame)
        self.rule_pattern.grid(row=0, column=1, sticky=tk.EW, padx=(0, 5))
        self.rule_pattern.insert(0, "prefix_$NAME_$ISEQ3")
        
        # 使用更紧凑的规则说明
        tip_text = "$NAME=原名 $ISEQ=序号 $ISEQ3=固定位序号(001) $YYYY=年 $MM=月 $DD=日"
        ttk.Label(rule_frame, text=tip_text, foreground="gray", font=("", 9)).grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(2, 5))
        
        # 序号设置 - 全部放在一行，更紧凑
        seq_frame = ttk.Frame(self.custom_rule_frame)
        seq_frame.pack(fill=tk.X, pady=5, padx=5)
        
        ttk.Label(seq_frame, text="起始值:").grid(row=0, column=0, sticky=tk.W, padx=2)
        self.start_value = ttk.Entry(seq_frame, width=4)
        self.start_value.grid(row=0, column=1, sticky=tk.W, padx=2)
        self.start_value.insert(0, "1")
        
        ttk.Label(seq_frame, text="步长:").grid(row=0, column=2, sticky=tk.W, padx=2)
        self.step = ttk.Entry(seq_frame, width=4)
        self.step.grid(row=0, column=3, sticky=tk.W, padx=2)
        self.step.insert(0, "1")
        
        ttk.Label(seq_frame, text="位数:").grid(row=0, column=4, sticky=tk.W, padx=2)
        self.digits = ttk.Entry(seq_frame, width=4)
        self.digits.grid(row=0, column=5, sticky=tk.W, padx=2)
        self.digits.insert(0, "3")
    
    def setup_output_settings(self, parent):
        """设置输出设置区域"""
        output_frame = ttk.LabelFrame(parent, text="输出设置")
        output_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=0)
        
        # 目标路径
        path_frame = ttk.Frame(output_frame)
        path_frame.pack(fill=tk.X, padx=10, pady=(10, 8))
        
        ttk.Label(path_frame, text="目标路径:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.target_path = tk.StringVar()
        ttk.Entry(path_frame, textvariable=self.target_path, width=30).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        browse_btn = ttk.Button(path_frame, text="浏览", command=self.browse_target_path, style="Auxiliary.TButton")
        browse_btn.pack(side=tk.LEFT, padx=0)
    
    def setup_preview_area(self, parent):
        """设置预览区域"""
        # 创建一个容器框架，用于设置预览表格和滚动条
        preview_container = ttk.Frame(parent)
        preview_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 预览表格
        columns = ("序号", "原始输入", "生成目录名", "完整路径")
        self.preview_tree = ttk.Treeview(preview_container, columns=columns, show="headings", height=5)
        
        # 设置列标题
        for col in columns:
            self.preview_tree.heading(col, text=col)
            if col == "序号":
                self.preview_tree.column(col, width=50, stretch=False)
            elif col == "原始输入":
                self.preview_tree.column(col, width=150)
            elif col == "生成目录名":
                self.preview_tree.column(col, width=180)
            else:  # 完整路径
                self.preview_tree.column(col, width=350, stretch=True)
        
        # 添加垂直滚动条
        y_scrollbar = ttk.Scrollbar(preview_container, orient="vertical", command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=y_scrollbar.set)
        
        # 添加水平滚动条
        x_scrollbar = ttk.Scrollbar(preview_container, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(xscrollcommand=x_scrollbar.set)
        
        # 布局，使用grid而不是pack以更好地控制布局
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        x_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # 配置容器的行列权重
        preview_container.grid_rowconfigure(0, weight=1)
        preview_container.grid_columnconfigure(0, weight=1)
    
    def toggle_input_method(self):
        """切换输入方式"""
        if self.input_method.get() == "direct":
            self.direct_input_frame.pack(fill=tk.X, padx=5, pady=5)
            self.file_input_frame.pack_forget()
        else:
            self.direct_input_frame.pack_forget()
            self.file_input_frame.pack(fill=tk.X, padx=5, pady=5)
    
    def toggle_naming_rule(self):
        """切换命名规则"""
        if self.naming_rule.get() == "direct":
            self.custom_rule_frame.pack_forget()
            self.direct_naming_frame.pack(fill=tk.BOTH, expand=True)
        else:
            self.direct_naming_frame.pack_forget()
            # 使用合适的布局参数
            self.custom_rule_frame.pack(fill=tk.BOTH, expand=True, pady=5)
            
            # 更新几何信息以获取实际尺寸
            self.custom_rule_frame.update_idletasks()
            
            # 确保父窗口调整大小以适应内容
            required_height = self.custom_rule_frame.winfo_reqheight()
            current_height = self.custom_rule_frame.master.winfo_height()
            
            if required_height > current_height:
                # 设置父框架的最小高度
                self.custom_rule_frame.master.configure(height=required_height + 10)  # 加10为边距
    
    def paste_from_clipboard(self):
        """从剪贴板粘贴"""
        try:
            text = self.root.clipboard_get()
            self.text_input.delete(1.0, tk.END)
            self.text_input.insert(1.0, text)
        except tk.TclError:
            messagebox.showerror("错误", "剪贴板为空或内容不可用")
    
    def clear_input(self):
        """清空输入"""
        self.text_input.delete(1.0, tk.END)
    
    def browse_file(self):
        """浏览文件"""
        self.logger.info("浏览文件")
        filetypes = [
            ("文本文件", "*.txt"),
            ("CSV文件", "*.csv"),
            ("Excel文件", "*.xlsx *.xls"),
            ("所有文件", "*.*")
        ]
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            self.file_path.set(filename)
            # 根据文件扩展名自动设置导入类型
            ext = os.path.splitext(filename)[1].lower()
            if ext in ['.txt', '.csv', '.xlsx', '.xls']:
                self.file_import_type.set(ext)
            # 尝试预览文件内容
            self.show_file_preview(filename)
            self.logger.info(f"已选择文件: {filename}")
    
    def browse_target_path(self):
        """浏览目标路径"""
        self.logger.info("浏览目标路径")
        path = filedialog.askdirectory()
        if path:
            self.target_path.set(path)
            self.logger.info(f"已选择目标路径: {path}")
    
    def preview(self):
        """预览生成结果"""
        self.logger.info("开始预览生成结果")
        try:
            # 清空预览表格
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
            
            # 获取输入内容
            dir_names = self.get_input_names()
            if not dir_names:
                self.logger.warning("没有输入目录名称")
                messagebox.showerror("错误", "请输入目录名称")
                return
            
            # 获取目标路径
            target_path = self.target_path.get()
            if not target_path:
                self.logger.warning("未选择目标路径")
                messagebox.showerror("错误", "请选择目标路径")
                return
            
            # 获取命名规则
            naming_rule = None
            if self.naming_rule.get() == "custom":
                naming_rule = self.rule_pattern.get()
                if not naming_rule:
                    self.logger.warning("未设置命名规则")
                    messagebox.showerror("错误", "请输入命名规则")
                    return
            
            # 获取序号设置
            try:
                start_value = int(self.start_value.get())
                step = int(self.step.get())
                digits = int(self.digits.get())
            except ValueError:
                self.logger.error("序号设置无效")
                messagebox.showerror("错误", "起始值、步长和位数必须是整数")
                return
            
            # 禁用层级结构设置
            enable_hierarchy = False
            
            # 获取日期信息
            now = datetime.now()
            
            # 普通预览，不处理层级
            for i, name in enumerate(dir_names):
                seq = start_value + i * step
                seq_str = str(seq).zfill(digits)
                
                if naming_rule:
                    new_name = naming_rule.replace('$NAME', name)
                    
                    # 处理序号，支持指定位数
                    seq_matches = re.findall(r'\$ISEQ(\d*)', new_name)
                    for seq_match in seq_matches:
                        if seq_match:
                            seq_digits = int(seq_match)
                            seq_formatted = str(seq).zfill(seq_digits)
                            new_name = new_name.replace(f'$ISEQ{seq_match}', seq_formatted)
                        else:
                            new_name = new_name.replace('$ISEQ', str(seq))
                    
                    # 处理日期变量
                    new_name = new_name.replace('$YYYY', now.strftime('%Y'))
                    new_name = new_name.replace('$MM', now.strftime('%m'))
                    new_name = new_name.replace('$DD', now.strftime('%d'))
                else:
                    new_name = name
                
                # 生成完整路径
                full_path = os.path.join(target_path, new_name)
                
                # 添加到预览表格，使用斑马条纹
                tags = ("oddrow",) if i % 2 == 1 else ()
                self.preview_tree.insert("", tk.END, values=(i+1, name, new_name, full_path), tags=tags)
            
            self.logger.info(f"预览完成，显示了{len(dir_names)}个目录")
        
        except Exception as e:
            self.logger.error(f"预览失败: {str(e)}")
            messagebox.showerror("错误", f"预览失败: {str(e)}")
    
    def execute(self):
        """执行创建操作"""
        self.logger.info("开始执行创建目录")
        try:
            # 获取输入内容
            dir_names = self.get_input_names()
            if not dir_names:
                self.logger.warning("没有输入目录名称")
                messagebox.showerror("错误", "请输入目录名称")
                return
            
            # 获取目标路径
            target_path = self.target_path.get()
            if not target_path:
                self.logger.warning("未选择目标路径")
                messagebox.showerror("错误", "请选择目标路径")
                return
            
            # 获取命名规则
            naming_rule = None
            if self.naming_rule.get() == "custom":
                naming_rule = self.rule_pattern.get()
                if not naming_rule:
                    self.logger.warning("未设置命名规则")
                    messagebox.showerror("错误", "请输入命名规则")
                    return
            
            # 获取序号设置
            try:
                start_value = int(self.start_value.get())
                step = int(self.step.get())
                digits = int(self.digits.get())
            except ValueError:
                self.logger.error("序号设置无效")
                messagebox.showerror("错误", "起始值、步长和位数必须是整数")
                return
            
            # 禁用层级结构
            enable_hierarchy = False
            indent_spaces = 4  # 默认值
            
            # 调用create_dirs函数创建目录
            self.logger.info(f"开始创建目录，共{len(dir_names)}个，目标路径: {target_path}")
            
            # 记录详细参数
            params = {
                "目录数量": len(dir_names),
                "目标路径": target_path,
                "启用层级": False,
                "命名规则": naming_rule if naming_rule else "直接命名",
                "起始序号": start_value,
                "序号步长": step,
                "序号位数": digits
            }
            self.logger.info(f"创建目录参数: {params}")
            
            success, message = create_dirs(
                dir_names=dir_names,
                parent_dir=target_path,
                structure=None,
                naming_rule=naming_rule,
                start_value=start_value,
                step=step,
                digits=digits,
                enable_hierarchy=False,
                indent_spaces=4
            )
            
            if success:
                self.logger.info(f"创建目录成功: {message}")
                messagebox.showinfo("成功", message)
            else:
                self.logger.error(f"创建目录失败: {message}")
                messagebox.showerror("错误", message)
                
        except Exception as e:
            self.logger.error(f"执行创建目录失败: {str(e)}")
            messagebox.showerror("错误", f"创建目录时发生错误: {str(e)}")
    
    def _format_cell_value(self, value):
        """格式化单元格值，处理数字格式
        
        将浮点数转换为整数(如果可能)，移除.0后缀
        """
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    
    def show_file_preview(self, file_path):
        """显示文件内容预览"""
        try:
            if not file_path:
                return
                
            ext = os.path.splitext(file_path)[1].lower()
            preview_text = ""
            
            # 获取是否跳过表头行的设置
            skip_header = self.file_has_header.get()
            
            if ext == '.txt':
                # 尝试不同编码读取前10行
                for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            lines = f.readlines()
                            if skip_header and lines:  # 如果需要跳过表头
                                lines = lines[1:]
                            # 最多显示10行
                            preview_lines = lines[:10]
                            preview_text = ''.join(preview_lines)
                        break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        self.logger.error(f"使用{encoding}编码读取文件失败: {str(e)}")
                
                if not preview_text:
                    preview_text = "无法预览文件内容：未能以支持的编码方式读取文件"
            
            elif ext == '.csv':
                # 尝试不同编码读取CSV
                for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            lines = f.readlines()
                            if skip_header and lines:  # 如果需要跳过表头
                                lines = lines[1:]
                            # 最多显示5行
                            preview_lines = lines[:5]
                            preview_text = ''.join(preview_lines)
                        break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        self.logger.error(f"使用{encoding}编码读取CSV文件失败: {str(e)}")
                
                if not preview_text:
                    preview_text = "无法预览文件内容：未能以支持的编码方式读取CSV文件"
            
            elif ext == '.xlsx':
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb.active
                    
                    # 读取前5行
                    preview_rows = []
                    start_row = 1
                    if skip_header:  # 如果需要跳过表头
                        start_row = 2
                    
                    max_rows = min(start_row + 4, ws.max_row + 1)  # 最多显示5行
                    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=max_rows)):
                        cells = [cell.value if cell.value is not None else '' for cell in row]
                        preview_rows.append(','.join(self._format_cell_value(c) for c in cells))
                    
                    preview_text = '\n'.join(preview_rows)
                    wb.close()
                except Exception as e:
                    preview_text = f"无法预览Excel文件: {str(e)}"
                    self.logger.error(f"预览XLSX文件失败: {str(e)}")
            
            elif ext == '.xls':
                try:
                    import xlrd
                    wb = xlrd.open_workbook(file_path)
                    ws = wb.sheet_by_index(0)  # 获取第一个工作表
                    
                    # 读取前5行
                    preview_rows = []
                    start_row = 0
                    if skip_header:  # 如果需要跳过表头
                        start_row = 1
                    
                    for i in range(start_row, min(start_row + 5, ws.nrows)):
                        cells = [ws.cell_value(i, j) for j in range(ws.ncols)]
                        preview_rows.append(','.join(self._format_cell_value(c) for c in cells))
                    
                    preview_text = '\n'.join(preview_rows)
                except Exception as e:
                    preview_text = f"无法预览Excel文件: {str(e)}"
                    self.logger.error(f"预览XLS文件失败: {str(e)}")
            
            # 更新预览文本
            self.file_preview.config(state="normal")
            self.file_preview.delete(1.0, tk.END)
            self.file_preview.insert(1.0, preview_text)
            self.file_preview.config(state="disabled")
            self.logger.info(f"已预览文件内容: {file_path}")
            
        except Exception as e:
            self.logger.error(f"预览文件内容失败: {str(e)}")
            self.file_preview.config(state="normal")
            self.file_preview.delete(1.0, tk.END)
            self.file_preview.insert(1.0, f"无法预览文件内容: {str(e)}")
            self.file_preview.config(state="disabled")
    
    def get_input_names(self):
        """获取输入的名称列表"""
        self.logger.info("获取输入名称列表")
        if self.input_method.get() == "direct":
            # 从文本输入获取
            content = self.text_input.get(1.0, tk.END).strip()
            names = [line.strip() for line in content.split('\n') if line.strip()]
            self.logger.info(f"从直接输入获取了{len(names)}个名称")
            return names
        else:
            # 从文件获取
            file_path = self.file_path.get()
            if not file_path:
                self.logger.warning("未选择文件")
                messagebox.showerror("错误", "请选择一个文件")
                return []
            
            # 获取文件导入类型
            import_type = self.file_import_type.get()
            names = []
            
            try:
                # 从UI获取列索引和表头设置
                column_idx = int(self.column_index.get()) - 1  # 转换为0基索引
                has_header = self.file_has_header.get()
                self.logger.info(f"从文件导入，文件类型: {import_type}, 列索引: {column_idx+1}, 跳过表头: {has_header}")
                
                if import_type == '.txt':
                    # 尝试不同编码读取文件
                    for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                        try:
                            with open(file_path, 'r', encoding=encoding) as f:
                                lines = f.readlines()
                                if has_header and lines:
                                    lines = lines[1:]  # 跳过表头
                                names = [line.strip() for line in lines if line.strip()]
                            # 成功读取，跳出循环
                            break
                        except UnicodeDecodeError:
                            continue
                        except Exception as e:
                            self.logger.error(f"使用{encoding}编码读取TXT文件失败: {str(e)}")
                    
                    if not names:
                        self.logger.error("无法读取文件，可能是编码问题")
                        messagebox.showerror("错误", "无法读取文件，请检查文件编码")
                        return []
                
                elif import_type == '.csv':
                    # 尝试不同编码读取CSV
                    import csv
                    for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                        try:
                            with open(file_path, 'r', encoding=encoding) as f:
                                reader = csv.reader(f)
                                rows = list(reader)  # 将所有行读入内存
                                
                                if has_header and rows:
                                    rows = rows[1:]  # 跳过表头
                                
                                for row in rows:
                                    if row and len(row) > column_idx:
                                        name = row[column_idx].strip()
                                        if name:
                                            names.append(name)
                            # 成功读取，跳出循环
                            break
                        except UnicodeDecodeError:
                            continue
                        except Exception as e:
                            self.logger.error(f"使用{encoding}编码读取CSV文件失败: {str(e)}")
                    
                    if not names:
                        self.logger.error("无法读取CSV文件")
                        messagebox.showerror("错误", "无法读取CSV文件，请检查文件编码")
                        return []
                
                elif import_type == '.xlsx':
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        ws = wb.active
                        start_row = 2 if has_header else 1  # 如果有表头，从第2行开始
                        for row in ws.iter_rows(min_row=start_row):
                            if len(row) > column_idx:
                                cell_value = row[column_idx].value
                                if cell_value is not None:
                                    formatted_value = self._format_cell_value(cell_value)
                                    if formatted_value.strip():
                                        names.append(formatted_value)
                        wb.close()
                    except ImportError:
                        self.logger.error("未安装openpyxl库，无法读取Excel文件")
                        messagebox.showerror("错误", "读取Excel文件需要安装openpyxl库")
                        return []
                    except Exception as e:
                        self.logger.error(f"读取Excel文件失败: {str(e)}")
                        messagebox.showerror("错误", f"读取Excel文件出错: {str(e)}")
                        return []
                
                elif import_type == '.xls':
                    try:
                        import xlrd
                        wb = xlrd.open_workbook(file_path)
                        ws = wb.sheet_by_index(0)  # 获取第一个工作表
                        
                        start_row = 1 if has_header else 0  # 如果有表头，从第2行开始
                        for i in range(start_row, ws.nrows):
                            if ws.ncols > column_idx:
                                cell_value = ws.cell_value(i, column_idx)
                                if cell_value is not None:
                                    formatted_value = self._format_cell_value(cell_value)
                                    if formatted_value.strip():
                                        names.append(formatted_value)
                    except ImportError:
                        self.logger.error("未安装xlrd库，无法读取旧版Excel文件")
                        messagebox.showerror("错误", "读取.xls格式Excel文件需要安装xlrd库")
                        return []
                    except Exception as e:
                        self.logger.error(f"读取Excel文件失败: {str(e)}")
                        messagebox.showerror("错误", f"读取Excel文件出错: {str(e)}")
                        return []
                
                self.logger.info(f"从文件{import_type}中读取了{len(names)}个名称")
                return names
                
            except Exception as e:
                self.logger.error(f"从文件获取名称失败: {str(e)}")
                messagebox.showerror("错误", f"读取文件失败: {str(e)}")
                return [] 

    def refresh_file_preview(self):
        """刷新文件预览"""
        self.logger.info("刷新文件预览")
        try:
            # 获取文件路径
            file_path = self.file_path.get()
            if not file_path:
                self.logger.warning("未选择文件")
                return
                
            # 预览文件内容
            self.show_file_preview(file_path)
        except Exception as e:
            self.logger.error(f"刷新文件预览失败: {str(e)}")
            self.file_preview.config(state="normal")
            self.file_preview.delete(1.0, tk.END)
            self.file_preview.insert(1.0, f"刷新文件预览失败: {str(e)}")
            self.file_preview.config(state="disabled")
    
    def destroy(self):
        """销毁标签页时解绑所有事件"""
        try:
            # 关闭所有额外资源
            self.logger.info("关闭标签页资源")
        except Exception as e:
            self.logger.error(f"关闭资源失败: {str(e)}")
        finally:
            # 调用父类的destroy方法
            super().destroy() 