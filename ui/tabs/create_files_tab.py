import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from utils.file_utils import create_files
import os
import csv
import logging

class CreateFilesTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.logger = logging.getLogger("create_files_tab")
        self.logger.info("初始化创建文件标签页")
        
        # 设置自身的背景色和边框
        self.configure(style="TabContent.TFrame")
        
        self.setup_ui()
        self.setup_drag_drop()
        
        # 添加一些默认内容，避免界面显示为空白
        self.text_input.insert("1.0", "# 请在此输入文件名称(每行一个)\nfile1\nfile2\nfile3")
        self.text_input.configure(background="#FFFFFF", font=('Arial', 10))
        
        # 设置默认目标路径
        default_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.target_path.set(default_path)
        
        # 设置内容模板示例
        self.content_template.insert("1.0", "这是一个示例文件\n文件名: ${NAME}\n序号: ${ISEQ}")
        
        self.logger.info("创建文件标签页初始化完成")
        
    def setup_ui(self):
        """设置用户界面"""
        # 主框架
        main_frame = ttk.Frame(self, padding=(10, 10))
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 上部分：文件输入和设置
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 设置左右布局
        left_pane = ttk.Frame(top_frame, width=400)  # 设置固定宽度
        left_pane.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        left_pane.pack_propagate(False)  # 防止子组件改变frame大小
        
        right_pane = ttk.Frame(top_frame, width=400)  # 设置固定宽度
        right_pane.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        right_pane.pack_propagate(False)  # 防止子组件改变frame大小
        
        # 设置各部分
        self.setup_input_area(left_pane)      # 左侧：输入区域
        self.setup_naming_rules(right_pane)   # 右上：命名规则
        self.setup_output_settings(right_pane) # 右下：输出设置
        
        # 预览区域 - 限制高度
        preview_frame = ttk.LabelFrame(main_frame, text="预览", height=150)
        preview_frame.pack(fill=tk.X, expand=False)
        preview_frame.pack_propagate(False)  # 防止子组件改变frame高度
        
        self.setup_preview_area(preview_frame)
    
    def setup_input_area(self, parent):
        """设置输入区域"""
        input_frame = ttk.LabelFrame(parent, text="输入")
        input_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 输入方式选择
        method_frame = ttk.Frame(input_frame)
        method_frame.pack(fill=tk.X, padx=10, pady=(8, 5))
        
        ttk.Label(method_frame, text="输入方式:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.input_method = tk.StringVar(value="direct")
        ttk.Radiobutton(method_frame, text="直接输入", variable=self.input_method, value="direct", 
                       command=self.toggle_input_method).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(method_frame, text="从文件导入", variable=self.input_method, value="file", 
                       command=self.toggle_input_method).pack(side=tk.LEFT, padx=0)
        
        # 直接输入框架
        self.direct_input_frame = ttk.Frame(input_frame)
        self.direct_input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 添加多行文本输入框，带滚动条
        self.text_input = scrolledtext.ScrolledText(self.direct_input_frame, height=10, width=40, wrap=tk.WORD)
        self.text_input.pack(fill=tk.BOTH, expand=True)
        
        # 设置右键菜单
        self.setup_context_menu()
        
        # 从文件导入框架
        self.file_input_frame = ttk.Frame(input_frame)
        
        file_select_frame = ttk.Frame(self.file_input_frame)
        file_select_frame.pack(fill=tk.X, pady=5)
        
        self.file_path = tk.StringVar()
        ttk.Entry(file_select_frame, textvariable=self.file_path, width=30).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(file_select_frame, text="浏览", command=self.browse_file, style="Auxiliary.TButton").pack(side=tk.LEFT, padx=(5, 0))
        
        # 文件类型选择
        file_type_frame = ttk.Frame(self.file_input_frame)
        file_type_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(file_type_frame, text="文件类型:").pack(side=tk.LEFT)
        self.file_import_type = tk.StringVar(value=".txt")
        file_type_combo = ttk.Combobox(file_type_frame, textvariable=self.file_import_type, width=15)
        file_type_combo['values'] = ('.txt', '.csv', '.xlsx', '.xls')
        file_type_combo.pack(side=tk.LEFT, padx=(5, 0))
        
        # 表头和列选择选项
        options_frame = ttk.Frame(self.file_input_frame)
        options_frame.pack(fill=tk.X, pady=5)
        
        self.file_has_header = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="跳过表头行", variable=self.file_has_header).pack(side=tk.LEFT)
        
        column_frame = ttk.Frame(options_frame)
        column_frame.pack(side=tk.LEFT, padx=(20, 0))
        
        ttk.Label(column_frame, text="列索引:").pack(side=tk.LEFT)
        self.column_index = tk.StringVar(value="1")
        ttk.Entry(column_frame, textvariable=self.column_index, width=5).pack(side=tk.LEFT, padx=(5, 0))
        
        # 文件预览区域
        ttk.Label(self.file_input_frame, text="文件内容预览:").pack(anchor=tk.W, pady=(5, 0))
        
        self.file_preview = scrolledtext.ScrolledText(self.file_input_frame, height=8, width=40, wrap=tk.WORD, state="disabled")
        self.file_preview.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 默认显示直接输入
        self.toggle_input_method()
        
        # 设置文件拖放支持
        self.setup_drag_drop()
    
    def setup_naming_rules(self, parent):
        """设置命名规则区域"""
        naming_frame = ttk.LabelFrame(parent, text="命名规则")
        naming_frame.pack(fill=tk.X, padx=5, pady=(0, 10))
        
        # 命名方式选择
        method_frame = ttk.Frame(naming_frame)
        method_frame.pack(fill=tk.X, padx=10, pady=(8, 5))
        
        ttk.Label(method_frame, text="命名方式:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.naming_rule = tk.StringVar(value="direct")
        ttk.Radiobutton(method_frame, text="直接命名", variable=self.naming_rule, value="direct", 
                       command=self.toggle_naming_rule).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(method_frame, text="自定义规则", variable=self.naming_rule, value="custom", 
                       command=self.toggle_naming_rule).pack(side=tk.LEFT, padx=0)
        
        # 直接命名框架（无需额外设置，直接使用输入的名称）
        self.direct_naming_frame = ttk.Frame(naming_frame)
        ttk.Label(self.direct_naming_frame, text="将直接使用输入的文件名").pack(pady=10)
        
        # 自定义命名规则框架
        self.custom_naming_frame = ttk.Frame(naming_frame)
        
        rule_entry_frame = ttk.Frame(self.custom_naming_frame)
        rule_entry_frame.pack(fill=tk.X, pady=(5, 8))
        
        ttk.Label(rule_entry_frame, text="命名规则:").pack(side=tk.LEFT, padx=(0, 5))
        self.rule_entry = ttk.Entry(rule_entry_frame)
        self.rule_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.rule_entry.insert(0, "prefix_$NAME_$ISEQ3")
        
        # 规则说明
        tip_frame = ttk.LabelFrame(self.custom_naming_frame, text="命名规则说明")
        tip_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        
        tips = """
$NAME - 替换为原始名称
$ISEQ - 替换为序号，如 1, 2, 3...
$ISEQ3 - 替换为固定位数序号，如 001, 002...
$YYYY - 替换为年份
$MM - 替换为月份
$DD - 替换为日期
        """
        ttk.Label(tip_frame, text=tips, justify=tk.LEFT).pack(padx=8, pady=8)
        
        # 序号设置
        seq_frame = ttk.Frame(self.custom_naming_frame)
        seq_frame.pack(fill=tk.X, pady=5)
        
        # 使用Grid布局管理序号设置，实现整齐的对齐
        ttk.Label(seq_frame, text="起始序号:").grid(row=0, column=0, sticky=tk.W, padx=(5, 5), pady=2)
        self.start_value = tk.StringVar(value="1")
        ttk.Entry(seq_frame, textvariable=self.start_value, width=5).grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)
        
        ttk.Label(seq_frame, text="序号步长:").grid(row=1, column=0, sticky=tk.W, padx=(5, 5), pady=2)
        self.step_value = tk.StringVar(value="1")
        ttk.Entry(seq_frame, textvariable=self.step_value, width=5).grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        
        # 默认显示直接命名
        self.toggle_naming_rule()
    
    def setup_output_settings(self, parent):
        """设置输出设置区域"""
        output_frame = ttk.LabelFrame(parent, text="输出设置")
        output_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=0)
        
        # 目标路径
        path_frame = ttk.Frame(output_frame)
        path_frame.pack(fill=tk.X, padx=10, pady=(10, 8))
        
        ttk.Label(path_frame, text="目标路径:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.target_path = tk.StringVar()
        ttk.Entry(path_frame, textvariable=self.target_path, width=40).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        browse_btn = ttk.Button(path_frame, text="浏览", command=self.browse_target_path, style="Auxiliary.TButton")
        browse_btn.pack(side=tk.LEFT, padx=0)
        
        # 文件类型
        type_frame = ttk.Frame(output_frame)
        type_frame.pack(fill=tk.X, padx=10, pady=8)
        
        ttk.Label(type_frame, text="文件类型:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.file_type = tk.StringVar(value=".txt")
        file_type_combo = ttk.Combobox(type_frame, textvariable=self.file_type, width=15)
        file_type_combo['values'] = ('.txt', '.md', '.html', '.css', '.js', '.py', '.json', '.xlsx', '.csv')
        file_type_combo.pack(side=tk.LEFT, padx=(0, 0))
        
        # 内容模板 - 使用分隔线增强视觉层次
        separator = ttk.Separator(output_frame, orient="horizontal")
        separator.pack(fill=tk.X, padx=5, pady=5)
        
        template_frame = ttk.LabelFrame(output_frame, text="内容模板 (可选)")
        template_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        self.content_template = scrolledtext.ScrolledText(template_frame, height=4, width=30, wrap=tk.WORD)
        self.content_template.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 模板说明
        ttk.Label(template_frame, text="使用 ${NAME} 替换为文件名, ${ISEQ} 替换为序号", 
                 foreground="gray").pack(anchor=tk.W, padx=5, pady=(0, 5))
    
    def setup_preview_area(self, parent):
        """设置预览区域"""
        # 创建一个容器框架，用于设置预览表格和滚动条
        preview_container = ttk.Frame(parent)
        preview_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 预览表格
        columns = ("序号", "原始名称", "文件名", "完整路径")
        self.preview_tree = ttk.Treeview(preview_container, columns=columns, show="headings", height=5)
        
        # 设置列标题
        for col in columns:
            self.preview_tree.heading(col, text=col)
            if col == "序号":
                self.preview_tree.column(col, width=50, stretch=False)
            elif col == "原始名称":
                self.preview_tree.column(col, width=150)
            elif col == "文件名":
                self.preview_tree.column(col, width=200)
            else:
                self.preview_tree.column(col, width=350)
        
        # 添加垂直滚动条
        y_scrollbar = ttk.Scrollbar(preview_container, orient="vertical", command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=y_scrollbar.set)
        
        # 添加水平滚动条
        x_scrollbar = ttk.Scrollbar(preview_container, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(xscrollcommand=x_scrollbar.set)
        
        # 布局，使用grid而不是pack以更好地控制布局
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        x_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # 配置容器的行列权重
        preview_container.grid_rowconfigure(0, weight=1)
        preview_container.grid_columnconfigure(0, weight=1)
    
    def setup_context_menu(self):
        """设置右键菜单"""
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="剪切", command=lambda: self.text_input.event_generate("<<Cut>>"))
        self.context_menu.add_command(label="复制", command=lambda: self.text_input.event_generate("<<Copy>>"))
        self.context_menu.add_command(label="粘贴", command=lambda: self.text_input.event_generate("<<Paste>>"))
        self.context_menu.add_separator()
        self.context_menu.add_command(label="全选", command=self.select_all)
        self.context_menu.add_command(label="清空", command=self.clear_input)
        
        self.text_input.bind("<Button-3>", self.show_context_menu)
    
    def setup_drag_drop(self):
        """设置拖放功能"""
        try:
            # 检查是否有TkinterDnD支持
            if hasattr(self.text_input, 'drop_target_register'):
                self.text_input.drop_target_register("DND_Files")
                self.text_input.dnd_bind("<<Drop>>", self.on_drop)
                self.logger.info("拖放功能设置成功")
                
                # 添加拖放提示
                hint_label = ttk.Label(
                    self.direct_input_frame, 
                    text="提示: 您可以直接拖放文件到此区域", 
                    foreground="gray",
                    background="#f5f5f5"
                )
                hint_label.pack(side=tk.BOTTOM, anchor=tk.W, padx=5, pady=(0, 5))
            else:
                self.logger.warning("没有找到拖放支持，请安装TkinterDnD2")
                
                # 添加安装建议
                hint_label = ttk.Label(
                    self.direct_input_frame, 
                    text="提示: 安装TkinterDnD2库以启用拖放功能", 
                    foreground="#FF6B6B",
                    background="#f5f5f5"
                )
                hint_label.pack(side=tk.BOTTOM, anchor=tk.W, padx=5, pady=(0, 5))
        except Exception as e:
            self.logger.warning(f"拖放功能设置失败: {str(e)}，可能需要安装TkinterDnD2")
    
    def on_drop(self, event):
        """处理拖放事件"""
        try:
            file_path = event.data
            # 移除引号和额外字符
            if file_path.startswith('{') and file_path.endswith('}'):
                file_path = file_path[1:-1]
            
            self.logger.info(f"接收到拖放文件: {file_path}")
            
            # 根据当前选择的输入方式处理
            if self.input_method.get() == "direct":
                self.read_file_content(file_path)
            else:
                self.file_path.set(file_path)
                # 根据文件扩展名自动设置导入类型
                ext = os.path.splitext(file_path)[1].lower()
                if ext in ['.txt', '.csv', '.xlsx', '.xls']:
                    self.file_import_type.set(ext)
                # 预览文件内容
                self.show_file_preview(file_path)
        except Exception as e:
            self.logger.error(f"处理拖放失败: {str(e)}")
            messagebox.showerror("错误", f"处理拖放失败: {str(e)}")
    
    def show_context_menu(self, event):
        """显示右键菜单"""
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
    
    def select_all(self):
        """全选文本"""
        self.text_input.tag_add(tk.SEL, "1.0", tk.END)
        self.text_input.mark_set(tk.INSERT, "1.0")
        self.text_input.see(tk.INSERT)
        return 'break'
    
    def toggle_input_method(self):
        """切换输入方式"""
        self.logger.debug(f"切换输入方式: {self.input_method.get()}")
        if self.input_method.get() == "direct":
            self.direct_input_frame.pack(fill=tk.X, padx=5, pady=5)
            self.file_input_frame.pack_forget()
        else:
            self.direct_input_frame.pack_forget()
            self.file_input_frame.pack(fill=tk.X, padx=5, pady=5)
    
    def toggle_naming_rule(self):
        """切换命名规则"""
        self.logger.debug(f"切换命名规则: {self.naming_rule.get()}")
        if self.naming_rule.get() == "direct":
            self.custom_naming_frame.pack_forget()
        else:
            self.custom_naming_frame.pack(fill=tk.X, padx=5, pady=5)
    
    def paste_from_clipboard(self):
        """从剪贴板粘贴"""
        try:
            self.logger.info("尝试从剪贴板粘贴")
            clipboard = self.clipboard_get()
            self.text_input.delete(1.0, tk.END)
            self.text_input.insert(1.0, clipboard)
            self.logger.info("粘贴成功，内容长度: {}".format(len(clipboard)))
        except Exception as e:
            self.logger.error(f"从剪贴板粘贴失败: {str(e)}")
            messagebox.showerror("错误", "剪贴板为空或内容不可用")
    
    def clear_input(self):
        """清空输入"""
        self.logger.debug("清空输入")
        self.text_input.delete(1.0, tk.END)
    
    def read_file_content(self, file_path):
        """读取文件内容"""
        self.logger.info(f"读取文件内容: {file_path}")
        ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if ext in ['.txt', '.md', '.csv', '.py', '.html', '.css', '.js']:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    self.text_input.delete(1.0, tk.END)
                    self.text_input.insert(1.0, content)
                    self.logger.info(f"成功读取文本文件，内容长度: {len(content)}")
                    
            elif ext in ['.xlsx', '.xls']:
                self.logger.info("检测到Excel文件，显示导入选项")
                self.input_method.set("file")
                self.toggle_input_method()
                self.file_path.set(file_path)
                self.file_import_type.set(ext)
                self.show_file_preview(file_path)
                
            else:
                self.logger.warning(f"不支持的文件类型: {ext}")
                messagebox.showwarning("警告", f"不支持的文件类型: {ext}")
                
        except Exception as e:
            self.logger.error(f"读取文件失败: {str(e)}")
            messagebox.showerror("错误", f"读取文件失败: {str(e)}")
    
    def browse_file(self):
        """浏览文件"""
        self.logger.info("浏览文件")
        filetypes = [
            ("文本文件", "*.txt"),
            ("CSV文件", "*.csv"),
            ("Excel文件", "*.xlsx *.xls"),
            ("所有文件", "*.*")
        ]
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            self.file_path.set(filename)
            # 根据文件扩展名自动设置导入类型
            ext = os.path.splitext(filename)[1].lower()
            if ext in ['.txt', '.csv', '.xlsx', '.xls']:
                self.file_import_type.set(ext)
            # 尝试预览文件内容
            self.show_file_preview(filename)
            self.logger.info(f"已选择文件: {filename}")
    
    def _format_cell_value(self, value):
        """格式化单元格值，处理数字格式
        
        将浮点数转换为整数(如果可能)，移除.0后缀
        """
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    
    def show_file_preview(self, file_path):
        """显示文件内容预览"""
        try:
            self.logger.info(f"预览文件内容: {file_path}")
            ext = os.path.splitext(file_path)[1].lower()
            preview_text = ""
            
            if ext == '.txt':
                # 尝试不同编码读取前10行
                for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            lines = [next(f, '') for _ in range(10)]
                        preview_text = ''.join(lines)
                        break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        self.logger.error(f"使用{encoding}编码读取文件失败: {str(e)}")
                
                if not preview_text:
                    preview_text = "无法预览文件内容：未能以支持的编码方式读取文件"
            
            elif ext == '.csv':
                # 尝试不同编码读取前5行
                for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            lines = [next(f, '') for _ in range(5)]
                        preview_text = ''.join(lines)
                        break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        self.logger.error(f"使用{encoding}编码读取CSV文件失败: {str(e)}")
                
                if not preview_text:
                    preview_text = "无法预览文件内容：未能以支持的编码方式读取CSV文件"
            
            elif ext == '.xlsx':
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb.active
                    
                    # 读取前5行
                    preview_rows = []
                    for i, row in enumerate(ws.iter_rows(max_row=5)):
                        cells = [cell.value if cell.value is not None else '' for cell in row]
                        preview_rows.append(','.join(self._format_cell_value(c) for c in cells))
                        if i >= 4:  # 最多5行
                            break
                    
                    preview_text = '\n'.join(preview_rows)
                    wb.close()
                except Exception as e:
                    preview_text = f"无法预览Excel文件: {str(e)}"
            
            elif ext == '.xls':
                try:
                    import xlrd
                    wb = xlrd.open_workbook(file_path)
                    ws = wb.sheet_by_index(0)  # 获取第一个工作表
                    
                    # 读取前5行
                    preview_rows = []
                    for i in range(min(5, ws.nrows)):
                        cells = [ws.cell_value(i, j) for j in range(ws.ncols)]
                        preview_rows.append(','.join(self._format_cell_value(c) for c in cells))
                    
                    preview_text = '\n'.join(preview_rows)
                except Exception as e:
                    preview_text = f"无法预览Excel文件: {str(e)}"
            
            # 更新预览文本
            self.file_preview.config(state="normal")
            self.file_preview.delete(1.0, tk.END)
            self.file_preview.insert(1.0, preview_text)
            self.file_preview.config(state="disabled")
            
        except Exception as e:
            self.logger.error(f"预览文件内容失败: {str(e)}")
            self.file_preview.config(state="normal")
            self.file_preview.delete(1.0, tk.END)
            self.file_preview.insert(1.0, f"无法预览文件内容: {str(e)}")
            self.file_preview.config(state="disabled")
    
    def browse_target_path(self):
        """浏览目标路径"""
        self.logger.info("浏览目标路径")
        path = filedialog.askdirectory()
        if path:
            self.target_path.set(path)
            self.logger.info(f"已选择目标路径: {path}")
    
    def get_input_names(self):
        """获取输入的名称列表"""
        self.logger.info("获取输入名称列表")
        if self.input_method.get() == "direct":
            # 从文本输入获取
            content = self.text_input.get(1.0, tk.END).strip()
            names = [line.strip() for line in content.split('\n') if line.strip()]
            self.logger.info(f"从直接输入获取了{len(names)}个名称")
            return names
        else:
            # 从文件获取
            file_path = self.file_path.get()
            if not file_path:
                self.logger.warning("未选择文件")
                messagebox.showerror("错误", "请选择一个文件")
                return []
            
            # 获取文件导入类型
            import_type = self.file_import_type.get()
            names = []
            
            try:
                # 从UI获取列索引和表头设置
                column_idx = int(self.column_index.get()) - 1  # 转换为0基索引
                has_header = self.file_has_header.get()
                
                if import_type == '.txt':
                    # 尝试不同编码读取文件
                    for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                        try:
                            with open(file_path, 'r', encoding=encoding) as f:
                                lines = f.readlines()
                                if has_header and lines:
                                    lines = lines[1:]  # 跳过表头
                                names = [line.strip() for line in lines if line.strip()]
                            # 成功读取，跳出循环
                            self.logger.info(f"使用{encoding}编码成功读取TXT文件")
                            break
                        except UnicodeDecodeError:
                            continue
                        except Exception as e:
                            self.logger.error(f"使用{encoding}编码读取TXT文件失败: {str(e)}")
                    
                    if not names:
                        self.logger.error("无法以任何支持的编码方式读取TXT文件")
                        messagebox.showerror("错误", "无法读取文件，请检查文件编码")
                        return []
                
                elif import_type == '.csv':
                    # 尝试不同编码读取CSV
                    for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                        try:
                            with open(file_path, 'r', encoding=encoding) as f:
                                reader = csv.reader(f)
                                rows = list(reader)  # 将所有行读入内存
                                
                                if has_header and rows:
                                    rows = rows[1:]  # 跳过表头
                                
                                for row in rows:
                                    if row and len(row) > column_idx:
                                        name = row[column_idx].strip()
                                        if name:
                                            names.append(name)
                            # 成功读取，跳出循环
                            self.logger.info(f"使用{encoding}编码成功读取CSV文件")
                            break
                        except UnicodeDecodeError:
                            continue
                        except Exception as e:
                            self.logger.error(f"使用{encoding}编码读取CSV文件失败: {str(e)}")
                    
                    if not names:
                        self.logger.error("无法以任何支持的编码方式读取CSV文件")
                        messagebox.showerror("错误", "无法读取CSV文件，请检查文件编码")
                        return []
                
                elif import_type == '.xlsx':
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        ws = wb.active
                        start_row = 2 if has_header else 1  # 如果有表头，从第2行开始
                        for row in ws.iter_rows(min_row=start_row):
                            if len(row) > column_idx:
                                cell_value = row[column_idx].value
                                if cell_value is not None:
                                    formatted_value = self._format_cell_value(cell_value)
                                    if formatted_value.strip():
                                        names.append(formatted_value)
                        wb.close()
                    except ImportError:
                        self.logger.error("未安装openpyxl库，无法读取Excel文件")
                        messagebox.showerror("错误", "读取Excel文件需要安装openpyxl库")
                        return []
                
                elif import_type == '.xls':
                    try:
                        import xlrd
                        wb = xlrd.open_workbook(file_path)
                        ws = wb.sheet_by_index(0)  # 获取第一个工作表
                        
                        start_row = 1 if has_header else 0  # 如果有表头，从第2行开始
                        for i in range(start_row, ws.nrows):
                            if ws.ncols > column_idx:
                                cell_value = ws.cell_value(i, column_idx)
                                if cell_value is not None:
                                    formatted_value = self._format_cell_value(cell_value)
                                    if formatted_value.strip():
                                        names.append(formatted_value)
                    except ImportError:
                        self.logger.error("未安装xlrd库，无法读取旧版Excel文件")
                        messagebox.showerror("错误", "读取.xls格式Excel文件需要安装xlrd库")
                        return []
                
                self.logger.info(f"从文件{import_type}中读取了{len(names)}个名称")
                return names
                
            except Exception as e:
                self.logger.error(f"从文件获取名称失败: {str(e)}")
                messagebox.showerror("错误", f"读取文件失败: {str(e)}")
                return []
    
    def preview(self):
        """预览生成结果"""
        self.logger.info("开始预览")
        try:
            # 清空预览表格
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
            
            # 获取输入名称
            names = self.get_input_names()
            if not names:
                self.logger.warning("没有输入名称")
                messagebox.showerror("错误", "请输入至少一个名称")
                return
            
            # 获取目标路径
            target_path = self.target_path.get()
            if not target_path:
                self.logger.warning("未选择目标路径")
                messagebox.showerror("错误", "请选择目标路径")
                return
            
            # 获取文件类型
            file_type = self.file_type.get()
            
            # 获取命名规则
            naming_rule = None
            if self.naming_rule.get() == "custom":
                naming_rule = self.rule_entry.get()
                self.logger.debug(f"使用自定义命名规则: {naming_rule}")
            
            # 获取序号设置
            try:
                start_value = int(self.start_value.get())
                step = int(self.step_value.get())
            except ValueError:
                self.logger.error("序号设置无效")
                messagebox.showerror("错误", "序号设置必须是整数")
                return
            
            # 生成预览
            from datetime import datetime
            import re
            
            for i, name in enumerate(names):
                seq = start_value + i * step
                seq_str = str(seq).zfill(3)
                
                if naming_rule:
                    new_name = naming_rule.replace('$NAME', name)
                    
                    # 处理序号，支持指定位数
                    seq_matches = re.findall(r'\$ISEQ(\d*)', new_name)
                    for seq_match in seq_matches:
                        if seq_match:
                            seq_digits = int(seq_match)
                            seq_formatted = str(seq).zfill(seq_digits)
                            new_name = new_name.replace(f'$ISEQ{seq_match}', seq_formatted)
                        else:
                            new_name = new_name.replace('$ISEQ', seq_str)
                    
                    # 处理日期变量
                    now = datetime.now()
                    new_name = new_name.replace('$YYYY', now.strftime('%Y'))
                    new_name = new_name.replace('$MM', now.strftime('%m'))
                    new_name = new_name.replace('$DD', now.strftime('%d'))
                else:
                    new_name = name
                
                # 添加文件类型扩展名
                filename = new_name + file_type
                full_path = os.path.join(target_path, filename)
                
                # 添加到预览表格，使用斑马条纹
                tags = ("oddrow",) if i % 2 == 1 else ()
                self.preview_tree.insert("", tk.END, values=(i+1, name, filename, full_path), tags=tags)
            
            self.logger.info(f"预览完成，显示了{len(names)}个文件")
        except Exception as e:
            self.logger.error(f"预览失败: {str(e)}")
            messagebox.showerror("错误", f"预览失败: {str(e)}")
    
    def execute(self):
        """执行创建文件"""
        self.logger.info("开始执行创建文件")
        try:
            # 获取输入名称
            names = self.get_input_names()
            if not names:
                self.logger.warning("没有输入名称")
                messagebox.showerror("错误", "请输入至少一个名称")
                return
            
            # 获取目标路径
            target_path = self.target_path.get()
            if not target_path:
                self.logger.warning("未选择目标路径")
                messagebox.showerror("错误", "请选择目标路径")
                return
            
            # 获取文件类型
            file_type = self.file_type.get()
            
            # 获取命名规则
            naming_rule = None
            if self.naming_rule.get() == "custom":
                naming_rule = self.rule_entry.get()
            
            # 获取序号设置
            try:
                start_value = int(self.start_value.get())
                step = int(self.step_value.get())
            except ValueError:
                self.logger.error("序号设置无效")
                messagebox.showerror("错误", "序号设置必须是整数")
                return
            
            # 获取内容模板
            content_template = self.content_template.get(1.0, tk.END).strip()
            if not content_template:
                content_template = None
            
            # 调用file_utils创建文件
            result, message = create_files(
                names=names,
                target_dir=target_path,
                file_type=file_type,
                content_template=content_template,
                naming_rule=naming_rule,
                start_value=start_value,
                step=step
            )
            
            if result:
                self.logger.info(f"创建文件成功: {message}")
                messagebox.showinfo("成功", message)
            else:
                self.logger.error(f"创建文件失败: {message}")
                messagebox.showerror("错误", message)
        except Exception as e:
            self.logger.error(f"执行创建文件失败: {str(e)}")
            messagebox.showerror("错误", f"执行失败: {str(e)}") 