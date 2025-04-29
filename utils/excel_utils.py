import os
import openpyxl
from datetime import datetime
from utils.log_utils import setup_logger, log_exception, log_operation_start, log_operation_end, log_file_operation, log_validation_result

# 使用统一的日志工具创建logger
logger = setup_logger('excel_utils')

def create_sheets(workbook_path, sheet_names, title_row=None, header_row=None):
    """
    批量创建工作表
    
    Args:
        workbook_path: Excel文件路径
        sheet_names: 工作表名称列表
        title_row: 标题行内容
        header_row: 表头行内容
    """
    log_operation_start(logger, "创建工作表", {
        "workbook_path": workbook_path,
        "sheet_count": len(sheet_names),
        "has_title": title_row is not None,
        "has_header": header_row is not None
    })
    
    success_count = 0
    fail_count = 0
    
    try:
        # 检查文件是否存在
        if os.path.exists(workbook_path):
            # 打开现有工作簿
            wb = openpyxl.load_workbook(workbook_path)
            logger.info(f"打开现有工作簿: {workbook_path}")
        else:
            # 创建新工作簿
            wb = openpyxl.Workbook()
            # 删除默认创建的工作表
            wb.remove(wb.active)
            logger.info(f"创建新工作簿: {workbook_path}")
            # 记录文件创建操作
            log_file_operation(logger, "创建", workbook_path)
        
        # 记录现有工作表
        logger.debug(f"当前工作簿中的工作表: {wb.sheetnames}")
        
        # 处理每个工作表名称
        for sheet_name in sheet_names:
            if not sheet_name.strip():
                logger.warning(f"跳过空白工作表名称")
                fail_count += 1
                continue
            
            # 验证工作表名称
            is_valid = True
            reason = None
            if len(sheet_name) > 31:  # Excel工作表名称限制
                is_valid = False
                reason = "工作表名称超过31个字符"
            elif any(c in sheet_name for c in [':', '\\', '/', '?', '*', '[', ']']):
                is_valid = False
                reason = "工作表名称包含非法字符"
                
            log_validation_result(logger, "工作表名称", sheet_name, is_valid, reason)
            
            if not is_valid:
                fail_count += 1
                continue
            
            # 检查工作表是否已存在
            if sheet_name in wb.sheetnames:
                logger.warning(f"工作表已存在: {sheet_name}")
                fail_count += 1
                continue
            
            # 创建工作表
            ws = wb.create_sheet(sheet_name)
            logger.info(f"创建工作表: {sheet_name}")
            
            # 添加标题行
            if title_row:
                ws['A1'] = title_row
                logger.debug(f"为工作表 {sheet_name} 添加标题行: {title_row}")
            
            # 添加表头行
            if header_row:
                for col, value in enumerate(header_row, 1):
                    ws.cell(row=2 if title_row else 1, column=col, value=value)
                logger.debug(f"为工作表 {sheet_name} 添加表头行: {header_row}")
            
            success_count += 1
        
        # 保存工作簿
        wb.save(workbook_path)
        logger.info(f"保存工作簿: {workbook_path}")
        
        log_operation_end(logger, "创建工作表", "成功", success_count, fail_count)
        return True, f"工作表创建成功: {success_count}个成功, {fail_count}个失败"
    
    except Exception as e:
        log_exception(logger, e, "创建工作表")
        log_operation_end(logger, "创建工作表", "失败", success_count, fail_count)
        return False, f"创建工作表失败: {str(e)}"

def read_sheet_names(file_path):
    """
    从Excel文件读取工作表名称
    
    Args:
        file_path: Excel文件路径
    """
    log_operation_start(logger, "读取工作表名称", {"file_path": file_path})
    
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            log_operation_end(logger, "读取工作表名称", "失败 - 文件不存在")
            return False, "文件不存在"
        
        # 验证文件类型
        _, ext = os.path.splitext(file_path)
        valid_extensions = ['.xlsx', '.xlsm', '.xltx', '.xltm']
        
        is_valid = ext.lower() in valid_extensions
        log_validation_result(
            logger, "Excel文件类型", ext, 
            is_valid, None if is_valid else f"不支持的文件类型，仅支持: {', '.join(valid_extensions)}"
        )
        
        if not is_valid:
            log_operation_end(logger, "读取工作表名称", "失败 - 不支持的文件类型")
            return False, f"不支持的文件类型，仅支持: {', '.join(valid_extensions)}"
        
        # 打开工作簿
        wb = openpyxl.load_workbook(file_path, read_only=True)
        logger.info(f"打开工作簿: {file_path}")
        
        # 获取工作表名称
        sheet_names = wb.sheetnames
        logger.info(f"读取到 {len(sheet_names)} 个工作表名称")
        logger.debug(f"工作表名称列表: {sheet_names}")
        
        log_operation_end(logger, "读取工作表名称", "成功", len(sheet_names))
        return True, sheet_names
    
    except Exception as e:
        log_exception(logger, e, "读取工作表名称")
        log_operation_end(logger, "读取工作表名称", "失败")
        return False, f"读取工作表名称失败: {str(e)}"

def read_column_data(file_path, sheet_name, column_index, has_header=True):
    """
    从Excel文件读取指定列的数据
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        column_index: 列索引（从1开始）
        has_header: 是否包含表头
    """
    log_operation_start(logger, "读取列数据", {
        "file_path": file_path,
        "sheet_name": sheet_name,
        "column_index": column_index,
        "has_header": has_header
    })
    
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            log_operation_end(logger, "读取列数据", "失败 - 文件不存在")
            return False, "文件不存在"
        
        # 打开工作簿
        wb = openpyxl.load_workbook(file_path, read_only=True)
        logger.info(f"打开工作簿: {file_path}")
        
        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            logger.error(f"工作表不存在: {sheet_name}")
            log_operation_end(logger, "读取列数据", "失败 - 工作表不存在")
            return False, "工作表不存在"
        
        # 获取工作表
        ws = wb[sheet_name]
        logger.debug(f"获取工作表: {sheet_name}")
        
        # 验证列索引
        max_column = ws.max_column
        is_valid = 1 <= column_index <= max_column
        log_validation_result(
            logger, "列索引", column_index, 
            is_valid, None if is_valid else f"列索引超出范围，有效范围: 1-{max_column}"
        )
        
        if not is_valid and max_column > 0:
            log_operation_end(logger, "读取列数据", "失败 - 列索引无效")
            return False, f"列索引无效，有效范围: 1-{max_column}"
        
        # 读取列数据
        data = []
        start_row = 2 if has_header else 1
        
        for row in ws.iter_rows(min_row=start_row, min_col=column_index, max_col=column_index):
            if row[0].value is not None:
                data.append(str(row[0].value))
        
        logger.info(f"从 {sheet_name} 工作表读取了 {len(data)} 行数据")
        logger.debug(f"读取的数据: {data[:10]}{'...' if len(data) > 10 else ''}")
        
        log_operation_end(logger, "读取列数据", "成功", len(data))
        return True, data
    
    except Exception as e:
        log_exception(logger, e, "读取列数据")
        log_operation_end(logger, "读取列数据", "失败")
        return False, f"读取列数据失败: {str(e)}" 