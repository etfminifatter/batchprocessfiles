import unittest
import os
import sys
import tempfile
import shutil
import logging

# 添加项目根目录到系统路径
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from utils.excel_utils import create_sheets, read_sheet_names, read_column_data

# 设置测试日志
logging.basicConfig(level=logging.ERROR)

class TestExcelUtils(unittest.TestCase):
    def setUp(self):
        # 创建临时目录用于测试
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, 'test_excel.xlsx')
    
    def tearDown(self):
        # 清理临时目录
        shutil.rmtree(self.temp_dir)
    
    def test_create_sheets_basic(self):
        """测试基本工作表创建功能"""
        sheet_names = ["Sheet1", "Sheet2", "测试表格"]
        
        result, message = create_sheets(
            workbook_path=self.test_file,
            sheet_names=sheet_names
        )
        
        self.assertTrue(result)
        
        # 验证创建的工作表
        success, result = read_sheet_names(self.test_file)
        self.assertTrue(success)
        self.assertEqual(set(sheet_names), set(result))
    
    def test_create_sheets_with_title(self):
        """测试带标题的工作表创建"""
        sheet_names = ["Sheet1", "Sheet2"]
        title_row = "测试标题"
        
        result, message = create_sheets(
            workbook_path=self.test_file,
            sheet_names=sheet_names,
            title_row=title_row
        )
        
        self.assertTrue(result)
        
        # 验证是否包含标题
        import openpyxl
        wb = openpyxl.load_workbook(self.test_file)
        for sheet_name in sheet_names:
            self.assertEqual(wb[sheet_name]['A1'].value, title_row)
    
    def test_create_sheets_with_header(self):
        """测试带表头的工作表创建"""
        sheet_names = ["Sheet1"]
        header_row = ["ID", "姓名", "年龄", "部门"]
        
        result, message = create_sheets(
            workbook_path=self.test_file,
            sheet_names=sheet_names,
            header_row=header_row
        )
        
        self.assertTrue(result)
        
        # 验证是否包含表头
        import openpyxl
        wb = openpyxl.load_workbook(self.test_file)
        ws = wb[sheet_names[0]]
        for i, header in enumerate(header_row, 1):
            self.assertEqual(ws.cell(row=1, column=i).value, header)
    
    def test_create_sheets_with_title_and_header(self):
        """测试同时带标题和表头的工作表创建"""
        sheet_names = ["Sheet1"]
        title_row = "测试标题"
        header_row = ["ID", "姓名", "年龄", "部门"]
        
        result, message = create_sheets(
            workbook_path=self.test_file,
            sheet_names=sheet_names,
            title_row=title_row,
            header_row=header_row
        )
        
        self.assertTrue(result)
        
        # 验证标题和表头
        import openpyxl
        wb = openpyxl.load_workbook(self.test_file)
        ws = wb[sheet_names[0]]
        
        # 验证标题
        self.assertEqual(ws['A1'].value, title_row)
        
        # 验证表头
        for i, header in enumerate(header_row, 1):
            self.assertEqual(ws.cell(row=2, column=i).value, header)
    
    def test_read_sheet_names(self):
        """测试读取工作表名称"""
        # 创建测试工作表
        sheet_names = ["Sheet1", "Sheet2", "Sheet3"]
        create_sheets(self.test_file, sheet_names)
        
        # 读取工作表名称
        success, result = read_sheet_names(self.test_file)
        
        self.assertTrue(success)
        self.assertEqual(set(sheet_names), set(result))
    
    def test_read_column_data(self):
        """测试读取列数据"""
        import openpyxl
        
        # 创建测试工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # 添加测试数据
        ws['A1'] = "ID"
        ws['B1'] = "姓名"
        ws['C1'] = "年龄"
        
        test_data = [
            [1, "张三", 25],
            [2, "李四", 30],
            [3, "王五", 28],
            [4, "赵六", 32],
            [5, "钱七", 29]
        ]
        
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        wb.save(self.test_file)
        
        # 读取姓名列数据（B列，索引为2）
        success, result = read_column_data(
            file_path=self.test_file,
            sheet_name="TestSheet",
            column_index=2,
            has_header=True
        )
        
        self.assertTrue(success)
        expected_names = ["张三", "李四", "王五", "赵六", "钱七"]
        self.assertEqual(expected_names, result)
    
    def test_read_column_data_no_header(self):
        """测试读取无表头的列数据"""
        import openpyxl
        
        # 创建测试工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # 添加测试数据（无表头）
        test_data = [
            [1, "张三", 25],
            [2, "李四", 30],
            [3, "王五", 28],
            [4, "赵六", 32],
            [5, "钱七", 29]
        ]
        
        for row_idx, row_data in enumerate(test_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        wb.save(self.test_file)
        
        # 读取姓名列数据（B列，索引为2）
        success, result = read_column_data(
            file_path=self.test_file,
            sheet_name="TestSheet",
            column_index=2,
            has_header=False
        )
        
        self.assertTrue(success)
        expected_names = ["张三", "李四", "王五", "赵六", "钱七"]
        self.assertEqual(expected_names, result)

if __name__ == "__main__":
    unittest.main() 